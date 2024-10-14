import pandas as pd
import openpyxl
import logging
import sys
from tkinter import messagebox

def prepare_data_sheet():
    """
    Load the CSV data into the Excel workbook's 'Data' sheet.
    """
    csv_file_path = "C:/Users/Inorw/Downloads/Data.csv"
    excel_file_path = "C:/Users/Inorw/OneDrive - University of Texas at San Antonio/Options.xlsx"

    try:
        # Load the CSV file into a pandas DataFrame
        data_df = pd.read_csv(csv_file_path)
        logging.info("SUCCESS. Loaded CSV data")

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        logging.info("SUCCESS. Loaded Excel workbook")

        # Access sheets in the workbook
        data_sheet = workbook['Data']
        main_sheet = workbook['Main']
        naked_put_sheet = workbook['Naked Put']

        # Clear existing data from the 'Data' sheet
        for row in data_sheet.iter_rows(min_row=1, min_col=1, max_row=data_sheet.max_row, max_col=data_sheet.max_column):
            for cell in row:
                cell.value = None
        logging.info("SUCCESS. Cleared existing data from the Data sheet")

        # Populate the 'Data' sheet with data from the DataFrame
        for row_index, row in enumerate(data_df.itertuples(index=False), start=1):
            for col_index, value in enumerate(row, start=1):
                data_sheet.cell(row=row_index, column=col_index, value=value)
        logging.info("SUCCESS. Populated Data sheet")

        # Save changes to the workbook
        workbook.save(excel_file_path)

        return data_sheet, main_sheet, naked_put_sheet, workbook

    except FileNotFoundError as e:
        logging.error(f"File not found: {str(e)}")
        messagebox.showerror("ERROR", f"File not found: {str(e)}")
        sys.exit()
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
        messagebox.showerror("ERROR", f"An unexpected error occurred: {str(e)}")
        sys.exit()
