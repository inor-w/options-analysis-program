import openpyxl
import os
import sys
import logging
import win32com.client
from tkinter import messagebox
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def close_workbook():
    """
    Close the Excel workbook if it's open to prevent conflicts.
    """
    excel_file_path = "C:/Users/Inorw/OneDrive - University of Texas at San Antonio/Options.xlsx"
    workbook_name = os.path.basename(excel_file_path)
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        workbook_found = False
        # Iterate over open workbooks in Excel
        for wb in excel.Workbooks:
            if os.path.basename(wb.FullName) == workbook_name:
                workbook_found = True
                wb.Close(SaveChanges=True)
                logging.info(f"Closed the workbook '{workbook_name}'")
        if not workbook_found:
            logging.info(f"The workbook '{workbook_name}' was not open")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
        messagebox.showerror("ERROR", f"An unexpected error occurred: {str(e)}")

def populate_sheets(calculation_results, data_sheet, main_sheet, naked_put_sheet, workbook):
    """
    Populate the 'Main' and 'Naked Put' sheets with the calculated data.
    """
    try:
        variables = calculation_results['variables']
        formatted_expiration = calculation_results['formatted_expiration']
        days_until_expiration = calculation_results['days_until_expiration']
        formatted_dateRetrieved = calculation_results['formatted_dateRetrieved']
        tradeTime = calculation_results['tradeTime']
        itm_probability = calculation_results['itm_probability']
        otm_probability = calculation_results['otm_probability']

        excel_file_path = "C:/Users/Inorw/OneDrive - University of Texas at San Antonio/Options.xlsx"

        typeVariable = None
        # Check if the option type is 'Cash Covered Put' to populate 'Naked Put' sheet
        if variables.get("Type") and "Cash Covered Put" in variables["Type"]:
            next_row = naked_put_sheet.max_row + 1
            typeVariable = "Cash Covered Put"

            # Mapping columns to variable names or calculated values for 'Naked Put' sheet
            nakedput_col_index_mapping = {
                1: "Account", 2: "Description", 3: "LastStockPrice", 4: formatted_expiration,
                5: "StrikePrice", 6: "BidOptionPrice", 7: "Volume", 8: "OpenInt", 9: "IVMid",
                10: "PurchasedPrice", 11: "MyOptionQuantity", 12: itm_probability, 13: "Delta",
                14: "PercentGainLoss", 15: "EarningsDate", 16: "Gamma", 17: "Vega", 18: "Theta",
                19: "Rho", 20: days_until_expiration, 21: formatted_dateRetrieved, 22: tradeTime,
                23: typeVariable
            }

            # Populate the 'Naked Put' sheet
            for col, var in nakedput_col_index_mapping.items():
                value = variables.get(var, var) if isinstance(var, str) else var
                naked_put_sheet.cell(row=next_row, column=col, value=value)

            # Format itm_probability as percentage
            naked_put_sheet.cell(row=next_row, column=12).number_format = '0.00%'

            # Apply conditional formatting for ITM column
            itm_range = f"L2:L{naked_put_sheet.max_row}"
            apply_conditional_formatting(naked_put_sheet, itm_range)

            # Save changes to the workbook
            workbook.save(excel_file_path)
            logging.info("SUCCESS. Populated Naked Put sheet")

        # Mapping columns to variable names or calculated values for 'Main' sheet
        main_col_index_mapping = {
            1: "Account", 2: "Description", 3: "LastStockPrice", 4: formatted_expiration,
            5: "StrikePrice", 6: "BidOptionPrice", 7: "Volume", 8: "OpenInt", 9: "IVMid",
            10: "PurchasedPrice", 11: "MyOptionQuantity", 12: itm_probability, 13: otm_probability,
            14: "Delta", 15: "PercentGainLoss", 16: "EarningsDate", 17: "Gamma", 18: "Vega",
            19: "Theta", 20: "Rho", 21: days_until_expiration, 22: formatted_dateRetrieved,
            23: tradeTime, 24: typeVariable
        }

        # Determine the next row to insert data in 'Main' sheet
        next_row = main_sheet.max_row + 1
        # Populate the 'Main' sheet
        for col, var in main_col_index_mapping.items():
            value = variables.get(var, var) if isinstance(var, str) else var
            main_sheet.cell(row=next_row, column=col, value=value)

        # Format probabilities as percentages
        main_sheet.cell(row=next_row, column=12).number_format = '0.00%'
        main_sheet.cell(row=next_row, column=13).number_format = '0.00%'

        # Apply conditional formatting for ITM probability in 'Main' sheet if applicable
        if variables.get("Type") and "Cash Covered Put" in variables["Type"]:
            itm_range = f"L2:L{main_sheet.max_row}"
            apply_conditional_formatting(main_sheet, itm_range)

        # Save changes to the workbook
        workbook.save(excel_file_path)
        logging.info("SUCCESS. Populated Main sheet")

        # Open the Excel file
        os.startfile(excel_file_path)

    except Exception as e:
        logging.error(f"An unexpected error occurred: {str(e)}")
        messagebox.showerror("ERROR", f"An unexpected error occurred: {str(e)}")
        sys.exit()

def apply_conditional_formatting(sheet, cell_range):
    """
    Apply conditional formatting to a given cell range in the sheet.
    Colors cells based on ITM probability thresholds.
    """
    # Define fill colors for different probability thresholds
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Define conditional formatting rules
    red_rule = CellIsRule(operator="lessThan", formula=["0.30"], fill=red_fill)  # Less than 30%
    yellow_rule = CellIsRule(operator="between", formula=["0.30", "0.50"], fill=yellow_fill)  # Between 30% and 50%
    green_rule = CellIsRule(operator="greaterThan", formula=["0.50"], fill=green_fill)  # Greater than 50%

    # Apply conditional formatting to the specified cell range
    sheet.conditional_formatting.add(cell_range, red_rule)
    sheet.conditional_formatting.add(cell_range, yellow_rule)
    sheet.conditional_formatting.add(cell_range, green_rule)
